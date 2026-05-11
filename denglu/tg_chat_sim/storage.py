import asyncio
import json
import re
from dataclasses import dataclass, asdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from redis.asyncio import Redis
from sqlalchemy import delete, select, update
from sqlalchemy.dialects.sqlite import insert as sqlite_insert
from sqlalchemy.exc import OperationalError as SAOperationalError
from sqlalchemy.ext.asyncio import AsyncEngine, AsyncSession, async_sessionmaker, create_async_engine
from sqlalchemy.pool import StaticPool

from tg_chat_sim.config import Settings, TelegramAccountConfig
from tg_chat_sim.models import (
    AccountProfileConfig,
    AccountProxyConfig,
    ChatScript,
    ChatScriptMessage,
    ChatCollectedMessage,
    ChatCollectProgress,
    ChatCollectState,
    Base,
    CopiedMessage,
)

try:
    from openpyxl import Workbook, load_workbook
except ModuleNotFoundError:
    Workbook = None  # type: ignore[misc, assignment]
    load_workbook = None  # type: ignore[misc, assignment]


def _sqlite_connect_failure_message(settings: Settings, exc: BaseException) -> str:
    orig = getattr(exc, "orig", exc)
    root_msg = str(orig).strip() or str(exc).strip()
    return (
        "无法打开 SQLite（当前 STORAGE_BACKEND 为 sqlite 或 both）。\n"
        f"- 文件: {settings.sqlite_path}\n"
        f"- 错误: {root_msg}\n\n"
        "请检查 SQLITE_DATABASE 路径是否可写、磁盘是否已满，或是否被其他进程独占。"
    )


@dataclass
class MessageRecord:
    source_group: str
    source_message_id: int
    sender_id: int | None
    sender_name: str | None
    text_content: str
    message_date: datetime


@dataclass
class RuntimeAccountRecord:
    session_name: str
    phone: str
    proxy_enabled: bool
    proxy_host: str
    proxy_port: int
    proxy_username: str
    proxy_password: str
    profile_username: str
    profile_user_id: int | None
    profile_avatar_local_path: str
    profile_bio: str
    profile_apply_changes: bool
    code_api_url: str = ""
    two_fa_password: str = ""


@dataclass
class ChatScriptMessageRecord:
    script_id: int
    speaker_index: int
    message_time: datetime
    text_content: str
    collected_message_id: int | None = None


@dataclass
class ChatCollectedMessageRecord:
    id: int
    text_content: str


@dataclass
class NamaProfileRecord:
    session_name: str
    username: str
    nickname: str
    avatar_local_path: str
    join_groups: list[str]
    apply_changes: bool


# 与 prepare_chat_records.py 导出的 xlsx 表头一致
_XLSX_HEADERS = [
    "msg_id",
    "msg_date",
    "sender_id",
    "sender_username",
    "sender_display_name",
    "text",
]


class StorageManager:
    def __init__(self, settings: Settings) -> None:
        self.settings = settings
        self.engine: AsyncEngine | None = None
        self.session_factory: async_sessionmaker[AsyncSession] | None = None
        self.redis_client: Redis | None = None
        self._account_runtime_path = Path(__file__).with_name("account_runtime.json")

    def _use_file_backend(self) -> bool:
        return self.settings.storage_backend == "file"

    def _use_sql_backend(self) -> bool:
        return self.settings.storage_backend in ("sqlite", "both")

    def _chat_records_path(self) -> Path:
        return self.settings.chat_records_path

    async def connect(self) -> None:
        if self._use_file_backend():
            self.engine = None
            self.session_factory = None
            self.redis_client = None
            return

        if self._use_sql_backend():
            self.engine = create_async_engine(
                self.settings.sqlite_url,
                echo=False,
                connect_args={"check_same_thread": False},
                poolclass=StaticPool,
            )
            self.session_factory = async_sessionmaker(self.engine, expire_on_commit=False)
            try:
                async with self.engine.begin() as conn:
                    await conn.run_sync(Base.metadata.create_all)
            except SAOperationalError as exc:
                await self.engine.dispose()
                self.engine = None
                self.session_factory = None
                raise RuntimeError(
                    _sqlite_connect_failure_message(self.settings, exc)
                ) from exc

        if self.settings.storage_backend in ("redis", "both"):
            self.redis_client = Redis.from_url(
                self.settings.redis_url,
                decode_responses=True,
                socket_connect_timeout=self.settings.redis_socket_connect_timeout,
                socket_timeout=self.settings.redis_socket_timeout,
            )
            await self.redis_client.ping()

    async def close(self) -> None:
        if self.redis_client:
            await self.redis_client.close()
        if self.engine:
            await self.engine.dispose()

    async def save_messages(self, messages: list[MessageRecord]) -> None:
        if not messages:
            return
        if self._use_file_backend():
            await asyncio.to_thread(self._append_messages_to_xlsx, messages)
            return
        if self._use_sql_backend():
            await self._save_to_sqlite(messages)
        if self.settings.storage_backend in ("redis", "both"):
            await self._save_to_redis(messages)

    async def load_messages(self, limit: int) -> list[MessageRecord]:
        if self._use_file_backend():
            return await asyncio.to_thread(self._load_messages_from_xlsx, limit)

        result: list[MessageRecord] = []
        if self._use_sql_backend() and self.session_factory:
            result = await self._load_from_sqlite(limit)
            if result:
                return result

        if self.settings.storage_backend in ("redis", "both") and self.redis_client:
            result = await self._load_from_redis(limit)
            if result:
                return result

        return result

    async def ensure_account_configs(
        self, accounts: list[TelegramAccountConfig], settings: Settings
    ) -> None:
        if self._use_file_backend():
            self._ensure_account_configs_file(accounts, settings)
            return
        if not self.session_factory:
            return

        proxy_rows = [
            {
                "session_name": acc.session_name,
                "enabled": settings.use_proxy,
                "proxy_host": settings.proxy_host,
                "proxy_port": settings.proxy_port,
                "proxy_username": settings.proxy_username,
                "proxy_password": settings.proxy_password,
                "updated_at": datetime.utcnow(),
            }
            for acc in accounts
        ]
        profile_rows = [
            {
                "session_name": acc.session_name,
                "username": "",
                "user_id": None,
                "avatar_local_path": "",
                "bio": "",
                "apply_changes": False,
                "updated_at": datetime.utcnow(),
            }
            for acc in accounts
        ]
        async with self.session_factory() as session:
            if proxy_rows:
                ins_p = sqlite_insert(AccountProxyConfig).values(proxy_rows)
                stmt_proxy = ins_p.on_conflict_do_update(
                    index_elements=[AccountProxyConfig.session_name],
                    set_={"updated_at": ins_p.excluded.updated_at},
                )
                await session.execute(stmt_proxy)
            if profile_rows:
                ins_pf = sqlite_insert(AccountProfileConfig).values(profile_rows)
                stmt_profile = ins_pf.on_conflict_do_update(
                    index_elements=[AccountProfileConfig.session_name],
                    set_={"updated_at": ins_pf.excluded.updated_at},
                )
                await session.execute(stmt_profile)
            await session.commit()

    async def load_runtime_accounts(
        self, accounts: list[TelegramAccountConfig], settings: Settings
    ) -> list[RuntimeAccountRecord]:
        if self._use_file_backend():
            return self._load_runtime_accounts_file(accounts, settings)

        if not self.session_factory:
            return [
                RuntimeAccountRecord(
                    session_name=acc.session_name,
                    phone=acc.phone,
                    proxy_enabled=settings.use_proxy,
                    proxy_host=settings.proxy_host,
                    proxy_port=settings.proxy_port,
                    proxy_username=settings.proxy_username,
                    proxy_password=settings.proxy_password,
                    profile_username="",
                    profile_user_id=None,
                    profile_avatar_local_path="",
                    profile_bio="",
                    profile_apply_changes=False,
                    code_api_url=getattr(acc, "code_api_url", "") or "",
                    two_fa_password=getattr(acc, "two_fa_password", "") or "",
                )
                for acc in accounts
            ]

        session_names = [acc.session_name for acc in accounts]
        async with self.session_factory() as session:
            proxy_rows = (
                await session.execute(
                    select(AccountProxyConfig).where(
                        AccountProxyConfig.session_name.in_(session_names)
                    )
                )
            ).scalars().all()
            profile_rows = (
                await session.execute(
                    select(AccountProfileConfig).where(
                        AccountProfileConfig.session_name.in_(session_names)
                    )
                )
            ).scalars().all()

        proxy_map = {row.session_name: row for row in proxy_rows}
        profile_map = {row.session_name: row for row in profile_rows}
        result: list[RuntimeAccountRecord] = []
        for acc in accounts:
            proxy = proxy_map.get(acc.session_name)
            profile = profile_map.get(acc.session_name)
            result.append(
                RuntimeAccountRecord(
                    session_name=acc.session_name,
                    phone=acc.phone,
                    proxy_enabled=proxy.enabled if proxy else settings.use_proxy,
                    proxy_host=proxy.proxy_host if proxy else settings.proxy_host,
                    proxy_port=proxy.proxy_port if proxy else settings.proxy_port,
                    proxy_username=(
                        proxy.proxy_username if proxy else settings.proxy_username
                    ),
                    proxy_password=(
                        proxy.proxy_password if proxy else settings.proxy_password
                    ),
                    profile_username=profile.username if profile else "",
                    profile_user_id=profile.user_id if profile else None,
                    profile_avatar_local_path=(
                        profile.avatar_local_path if profile else ""
                    ),
                    profile_bio=profile.bio if profile else "",
                    profile_apply_changes=profile.apply_changes if profile else False,
                    code_api_url=getattr(acc, "code_api_url", "") or "",
                    two_fa_password=getattr(acc, "two_fa_password", "") or "",
                )
            )
        return result

    async def sync_profile_snapshot(
        self,
        session_name: str,
        user_id: int | None,
        username: str | None,
    ) -> None:
        if self._use_file_backend():
            self._sync_profile_snapshot_file(session_name, user_id, username)
            return
        if not self.session_factory:
            return
        async with self.session_factory() as session:
            ins = sqlite_insert(AccountProfileConfig).values(
                [
                    {
                        "session_name": session_name,
                        "user_id": user_id,
                        "username": username or "",
                        "avatar_local_path": "",
                        "bio": "",
                        "apply_changes": False,
                        "updated_at": datetime.utcnow(),
                    }
                ]
            )
            stmt = ins.on_conflict_do_update(
                index_elements=[AccountProfileConfig.session_name],
                set_={
                    "user_id": ins.excluded.user_id,
                    "username": ins.excluded.username,
                    "updated_at": ins.excluded.updated_at,
                },
            )
            await session.execute(stmt)
            await session.commit()

    async def load_active_chat_script_messages(
        self, target_group: str, limit: int | None = None
    ) -> list[ChatScriptMessageRecord]:
        """
        file 后端：从 chat_records.xlsx 读取文本，按行分配给 10 个 speaker，时间轴用固定间隔生成。
        sqlite 后端：读库；若无 active 脚本或脚本无行，回退到同路径 xlsx（便于只部署 Excel 的场景）。
        """
        if self._use_file_backend():
            return await asyncio.to_thread(
                self._load_script_messages_from_xlsx, limit
            )

        if not self.session_factory:
            return await asyncio.to_thread(
                self._load_script_messages_from_xlsx, limit
            )

        async with self.session_factory() as session:
            script_stmt = (
                select(ChatScript.id)
                .where(ChatScript.target_group == target_group)
                .where(ChatScript.active == True)  # noqa: E712
                .order_by(ChatScript.created_at.desc())
                .limit(1)
            )
            script_id_row = (await session.execute(script_stmt)).first()
            if not script_id_row:
                return await asyncio.to_thread(
                    self._load_script_messages_from_xlsx, limit
                )
            script_id = int(script_id_row[0])

            msg_stmt = (
                select(ChatScriptMessage)
                .where(ChatScriptMessage.script_id == script_id)
                .order_by(ChatScriptMessage.message_time.asc())
            )
            if limit is not None:
                msg_stmt = msg_stmt.limit(limit)

            rows = (await session.execute(msg_stmt)).scalars().all()
            if rows:
                return [
                    ChatScriptMessageRecord(
                        script_id=script_id,
                        speaker_index=row.speaker_index,
                        message_time=row.message_time,
                        text_content=row.text_content,
                    )
                    for row in rows
                ]

        return await asyncio.to_thread(self._load_script_messages_from_xlsx, limit)

    async def reserve_unsent_collected_messages(
        self, limit: int | None = None
    ) -> list[ChatCollectedMessageRecord]:
        if self._use_file_backend() or not self.session_factory:
            return []
        async with self.session_factory() as session:
            stmt = (
                select(ChatCollectedMessage)
                .where(ChatCollectedMessage.used_flag == False)  # noqa: E712
                .order_by(ChatCollectedMessage.message_date.desc())
            )
            if limit is not None:
                stmt = stmt.limit(limit)
            rows = (await session.execute(stmt)).scalars().all()
            if not rows:
                return []
            ids = [int(r.id) for r in rows]
            await session.execute(
                update(ChatCollectedMessage)
                .where(ChatCollectedMessage.id.in_(ids))
                .values(used_flag=True)
            )
            await session.commit()
            return [
                ChatCollectedMessageRecord(id=int(r.id), text_content=r.text_content)
                for r in rows
            ]

    async def delete_collected_messages_by_ids(self, ids: list[int]) -> None:
        if not ids or self._use_file_backend() or not self.session_factory:
            return
        async with self.session_factory() as session:
            await session.execute(
                delete(ChatCollectedMessage).where(ChatCollectedMessage.id.in_(ids))
            )
            await session.commit()

    async def mark_collected_messages_unused(self, ids: list[int]) -> None:
        if not ids or self._use_file_backend() or not self.session_factory:
            return
        async with self.session_factory() as session:
            await session.execute(
                update(ChatCollectedMessage)
                .where(ChatCollectedMessage.id.in_(ids))
                .values(used_flag=False)
            )
            await session.commit()

    async def upsert_collected_messages(
        self,
        *,
        account_session_name: str,
        source_group: str,
        rows: list[dict],
    ) -> tuple[int, int]:
        if not rows or self._use_file_backend() or not self.session_factory:
            return (0, 0)
        payload = [
            {
                "account_session_name": account_session_name,
                "source_group": source_group,
                "source_message_id": int(r["source_message_id"]),
                "message_date": r["message_date"],
                "sender_id": r.get("sender_id"),
                "sender_username": str(r.get("sender_username") or ""),
                "sender_display_name": str(r.get("sender_display_name") or ""),
                "text_content": str(r["text_content"]),
                "used_flag": False,
                "created_at": datetime.utcnow(),
            }
            for r in rows
        ]
        source_ids = sorted(set(int(r["source_message_id"]) for r in rows))
        existed_before = await self.count_collected_messages_by_ids(
            source_group=source_group,
            source_ids=source_ids,
        )
        ins_c = sqlite_insert(ChatCollectedMessage).values(payload)
        stmt = ins_c.on_conflict_do_update(
            index_elements=[
                ChatCollectedMessage.source_group,
                ChatCollectedMessage.source_message_id,
            ],
            set_={
                "sender_id": ins_c.excluded.sender_id,
                "sender_username": ins_c.excluded.sender_username,
                "sender_display_name": ins_c.excluded.sender_display_name,
                "text_content": ins_c.excluded.text_content,
                "message_date": ins_c.excluded.message_date,
            },
        )
        async with self.session_factory() as session:
            await session.execute(stmt)
            await session.commit()
        existing_after = await self.count_collected_messages_by_ids(
            source_group=source_group,
            source_ids=source_ids,
        )
        inserted = max(0, existing_after - existed_before)
        return (inserted, len(payload))

    async def count_collected_messages_by_ids(
        self, *, source_group: str, source_ids: list[int]
    ) -> int:
        if not source_ids or self._use_file_backend() or not self.session_factory:
            return 0
        unique_ids = sorted(set(source_ids))
        async with self.session_factory() as session:
            stmt = select(ChatCollectedMessage.id).where(
                ChatCollectedMessage.source_group == source_group,
                ChatCollectedMessage.source_message_id.in_(unique_ids),
            )
            rows = (await session.execute(stmt)).all()
            return len(rows)

    async def get_collect_state(
        self, *, account_session_name: str, source_group: str
    ) -> dict:
        if self._use_file_backend() or not self.session_factory:
            return {
                "oldest_scanned_msg_id": 0,
                "last_collect_date": None,
                "daily_collected_count": 0,
            }
        async with self.session_factory() as session:
            stmt = (
                select(ChatCollectState)
                .where(ChatCollectState.account_session_name == account_session_name)
                .where(ChatCollectState.source_group == source_group)
                .limit(1)
            )
            row = (await session.execute(stmt)).scalars().first()
            if not row:
                return {
                    "oldest_scanned_msg_id": 0,
                    "last_collect_date": None,
                    "daily_collected_count": 0,
                }
            return {
                "oldest_scanned_msg_id": int(row.oldest_scanned_msg_id or 0),
                "last_collect_date": row.last_collect_date,
                "daily_collected_count": int(row.daily_collected_count or 0),
            }

    async def upsert_collect_state(
        self,
        *,
        account_session_name: str,
        source_group: str,
        oldest_scanned_msg_id: int | None = None,
        last_collect_date: date | None = None,
        daily_collected_count: int | None = None,
    ) -> None:
        if self._use_file_backend() or not self.session_factory:
            return
        values = {
            "account_session_name": account_session_name,
            "source_group": source_group,
            "oldest_scanned_msg_id": int(oldest_scanned_msg_id or 0),
            "last_collect_date": last_collect_date,
            "daily_collected_count": int(daily_collected_count or 0),
            "updated_at": datetime.utcnow(),
        }
        ins_s = sqlite_insert(ChatCollectState).values([values])
        update_values: dict[str, Any] = {"updated_at": ins_s.excluded.updated_at}
        if oldest_scanned_msg_id is not None:
            update_values["oldest_scanned_msg_id"] = ins_s.excluded.oldest_scanned_msg_id
        if last_collect_date is not None:
            update_values["last_collect_date"] = ins_s.excluded.last_collect_date
        if daily_collected_count is not None:
            update_values["daily_collected_count"] = ins_s.excluded.daily_collected_count
        stmt = ins_s.on_conflict_do_update(
            index_elements=[
                ChatCollectState.account_session_name,
                ChatCollectState.source_group,
            ],
            set_=update_values,
        )
        async with self.session_factory() as session:
            await session.execute(stmt)
            await session.commit()

    async def has_collect_progress(
        self, *, account_session_name: str, source_group: str, collect_date: date
    ) -> bool:
        if self._use_file_backend() or not self.session_factory:
            return False
        async with self.session_factory() as session:
            stmt = (
                select(ChatCollectProgress.id)
                .where(ChatCollectProgress.account_session_name == account_session_name)
                .where(ChatCollectProgress.source_group == source_group)
                .where(ChatCollectProgress.collect_date == collect_date)
                .limit(1)
            )
            return (await session.execute(stmt)).first() is not None

    async def add_collect_progress(
        self, *, account_session_name: str, source_group: str, collect_date: date
    ) -> None:
        if self._use_file_backend() or not self.session_factory:
            return
        ins_pr = sqlite_insert(ChatCollectProgress).values(
            [
                {
                    "account_session_name": account_session_name,
                    "source_group": source_group,
                    "collect_date": collect_date,
                    "status": "done",
                    "created_at": datetime.utcnow(),
                }
            ]
        )
        stmt = ins_pr.on_conflict_do_update(
            index_elements=[
                ChatCollectProgress.account_session_name,
                ChatCollectProgress.source_group,
                ChatCollectProgress.collect_date,
            ],
            set_={"status": ins_pr.excluded.status},
        )
        async with self.session_factory() as session:
            await session.execute(stmt)
            await session.commit()

    def load_nama_profiles(
        self, nama_path: str | None = None
    ) -> dict[str, NamaProfileRecord]:
        path = Path(nama_path) if nama_path else Path(__file__).with_name("nama.json")
        if not path.exists():
            return {}

        raw_text = path.read_text(encoding="utf-8")
        sanitized = re.sub(r",\s*([}\]])", r"\1", raw_text)
        payload = json.loads(sanitized) if sanitized.strip() else {}
        result: dict[str, NamaProfileRecord] = {}
        for session_name, info in payload.items():
            if not isinstance(info, dict):
                continue
            join_groups = info.get("join_groups") or []
            if isinstance(join_groups, str):
                raw_groups = join_groups.strip()
                if raw_groups.startswith("[") and raw_groups.endswith("]"):
                    inner = raw_groups[1:-1].strip()
                else:
                    inner = raw_groups
                join_groups = [
                    x.strip().strip("\"'").strip()
                    for x in re.split(r"[，,]", inner)
                    if x.strip().strip("\"'").strip()
                ]
            result[session_name] = NamaProfileRecord(
                session_name=session_name,
                username=str(info.get("username") or "").strip().lstrip("@"),
                nickname=str(info.get("nickname") or "").strip(),
                avatar_local_path=str(info.get("avatar_local_path") or "").strip(),
                join_groups=[str(x).strip() for x in join_groups if str(x).strip()],
                apply_changes=bool(int(info.get("apply_changes", 0))),
            )
        return result

    async def create_and_activate_chat_script(
        self, target_group: str
    ) -> int | None:
        if self._use_file_backend():
            return None
        assert self.session_factory is not None
        async with self.session_factory() as session:
            await session.execute(
                update(ChatScript).where(ChatScript.target_group == target_group).values(
                    active=False
                )
            )
            script = ChatScript(target_group=target_group, active=True)
            session.add(script)
            await session.flush()
            await session.commit()
            return int(script.id)

    async def save_chat_script_messages(
        self,
        script_id: int,
        messages_by_speaker: list[list[tuple[int, datetime, str]]],
    ) -> None:
        if self._use_file_backend():
            return
        assert self.session_factory is not None
        rows: list[dict] = []
        for speaker_list in messages_by_speaker:
            if not speaker_list:
                continue
            speaker_index = speaker_list[0][0]
            speaker_list_sorted = sorted(speaker_list, key=lambda x: x[1])
            for seq, (_, message_time, text_content) in enumerate(
                speaker_list_sorted, start=1
            ):
                rows.append(
                    {
                        "script_id": script_id,
                        "speaker_index": speaker_index,
                        "seq_in_speaker": seq,
                        "message_time": message_time,
                        "text_content": text_content,
                    }
                )

        if not rows:
            return

        stmt = sqlite_insert(ChatScriptMessage).values(rows)
        async with self.session_factory() as session:
            await session.execute(stmt)
            await session.commit()

    # --- file backend (chat_records.xlsx + account_runtime.json) ---

    def _load_account_runtime_raw(self) -> dict[str, Any]:
        if not self._account_runtime_path.exists():
            return {}
        try:
            data = json.loads(self._account_runtime_path.read_text(encoding="utf-8"))
            return data if isinstance(data, dict) else {}
        except Exception:
            return {}

    def _save_account_runtime_raw(self, data: dict[str, Any]) -> None:
        self._account_runtime_path.write_text(
            json.dumps(data, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def _ensure_account_configs_file(
        self, accounts: list[TelegramAccountConfig], settings: Settings
    ) -> None:
        data = self._load_account_runtime_raw()
        changed = False
        for acc in accounts:
            if acc.session_name in data:
                continue
            data[acc.session_name] = {
                "proxy_enabled": settings.use_proxy,
                "proxy_host": settings.proxy_host,
                "proxy_port": settings.proxy_port,
                "proxy_username": settings.proxy_username,
                "proxy_password": settings.proxy_password,
                "username": "",
                "user_id": None,
                "avatar_local_path": "",
                "bio": "",
                "apply_changes": False,
                "updated_at": datetime.utcnow().isoformat(),
            }
            changed = True
        if changed:
            self._save_account_runtime_raw(data)

    def _load_runtime_accounts_file(
        self, accounts: list[TelegramAccountConfig], settings: Settings
    ) -> list[RuntimeAccountRecord]:
        data = self._load_account_runtime_raw()
        result: list[RuntimeAccountRecord] = []
        for acc in accounts:
            row = data.get(acc.session_name, {})
            result.append(
                RuntimeAccountRecord(
                    session_name=acc.session_name,
                    phone=acc.phone,
                    proxy_enabled=bool(row.get("proxy_enabled", settings.use_proxy)),
                    proxy_host=str(row.get("proxy_host", settings.proxy_host)),
                    proxy_port=int(row.get("proxy_port", settings.proxy_port)),
                    proxy_username=str(
                        row.get("proxy_username", settings.proxy_username)
                    ),
                    proxy_password=str(
                        row.get("proxy_password", settings.proxy_password)
                    ),
                    profile_username=str(row.get("username", "")),
                    profile_user_id=row.get("user_id"),
                    profile_avatar_local_path=str(row.get("avatar_local_path", "")),
                    profile_bio=str(row.get("bio", "")),
                    profile_apply_changes=bool(row.get("apply_changes", False)),
                    code_api_url=getattr(acc, "code_api_url", "") or "",
                    two_fa_password=getattr(acc, "two_fa_password", "") or "",
                )
            )
        return result

    def _sync_profile_snapshot_file(
        self, session_name: str, user_id: int | None, username: str | None
    ) -> None:
        data = self._load_account_runtime_raw()
        if session_name not in data:
            data[session_name] = {
                "proxy_enabled": self.settings.use_proxy,
                "proxy_host": self.settings.proxy_host,
                "proxy_port": self.settings.proxy_port,
                "proxy_username": self.settings.proxy_username,
                "proxy_password": self.settings.proxy_password,
                "username": "",
                "user_id": None,
                "avatar_local_path": "",
                "bio": "",
                "apply_changes": False,
                "updated_at": datetime.utcnow().isoformat(),
            }
        data[session_name]["user_id"] = user_id
        data[session_name]["username"] = username or ""
        data[session_name]["updated_at"] = datetime.utcnow().isoformat()
        self._save_account_runtime_raw(data)

    def _require_openpyxl(self) -> None:
        if Workbook is None or load_workbook is None:
            raise ModuleNotFoundError(
                "缺少 openpyxl，请执行：pip install openpyxl"
            )

    def _script_xlsx_row_cap(self, limit: int | None) -> int:
        """聊天脚本从 xlsx 读取的最大行数（默认用 CHAT_SCRIPT_XLSX_MESSAGE_LIMIT）。"""
        if limit is not None:
            return max(1, limit)
        cap = getattr(
            self.settings, "chat_script_xlsx_message_limit", 2000
        )
        return max(1, min(int(cap), 50_000))

    @staticmethod
    def _coerce_xlsx_msg_id(raw: object, fallback: int) -> int:
        """Excel 常把整型 id 存成 float 或 '12345.0' 字符串，避免 int() 失败整行被丢弃。"""
        if raw is None:
            return fallback
        if isinstance(raw, bool):
            return fallback
        if isinstance(raw, int):
            return raw
        if isinstance(raw, float):
            if raw != raw:  # NaN
                return fallback
            return int(raw)
        s = str(raw).strip()
        if not s:
            return fallback
        try:
            return int(float(s))
        except (TypeError, ValueError):
            return fallback

    def _parse_xlsx_date(self, value: object) -> datetime:
        if isinstance(value, datetime):
            return value.replace(tzinfo=None) if value.tzinfo else value
        if value is None:
            return datetime.utcnow()
        s = str(value).strip()
        if not s:
            return datetime.utcnow()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(s.replace("Z", "+00:00")).replace(
                tzinfo=None
            )
        except ValueError:
            return datetime.utcnow()

    def _load_messages_from_xlsx(self, limit: int) -> list[MessageRecord]:
        self._require_openpyxl()
        path = self._chat_records_path()
        if not path.exists():
            return []

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows_iter = ws.iter_rows(min_row=1, values_only=True)
            header_row = next(rows_iter, None)
            if not header_row:
                return []
            header = [str(c).strip() if c is not None else "" for c in header_row]
            if header[: len(_XLSX_HEADERS)] != _XLSX_HEADERS:
                # 兼容：若表头不一致，按列位置读取
                pass

            idx = {name: i for i, name in enumerate(header) if name}
            def col(name: str, fallback: int) -> int:
                return idx.get(name, fallback)

            i_msg = col("msg_id", 0)
            i_date = col("msg_date", 1)
            i_sid = col("sender_id", 2)
            i_suser = col("sender_username", 3)
            i_sname = col("sender_display_name", 4)
            i_text = col("text", 5)

            def _cell(t: tuple[Any, ...] | list[Any], idx: int) -> object:
                return t[idx] if idx < len(t) else None

            out: list[MessageRecord] = []
            sg = str(self.settings.source_group)
            row_seq = 0
            for tup in rows_iter:
                if len(out) >= limit:
                    break
                if not tup or all(v is None or str(v).strip() == "" for v in tup):
                    continue
                row_seq += 1
                text = str(_cell(tup, i_text) or "").strip()
                if not text:
                    continue
                msg_id = self._coerce_xlsx_msg_id(_cell(tup, i_msg), fallback=row_seq)
                sender_name = (
                    str(_cell(tup, i_sname) or _cell(tup, i_suser) or "").strip()
                    or None
                )
                sid_raw = _cell(tup, i_sid)
                try:
                    sender_id = int(sid_raw) if sid_raw is not None else None
                except (TypeError, ValueError):
                    sender_id = None
                msg_dt = self._parse_xlsx_date(_cell(tup, i_date))
                out.append(
                    MessageRecord(
                        source_group=sg,
                        source_message_id=msg_id,
                        sender_id=sender_id,
                        sender_name=sender_name,
                        text_content=text,
                        message_date=msg_dt,
                    )
                )
            return out
        finally:
            wb.close()

    def _append_messages_to_xlsx(self, messages: list[MessageRecord]) -> None:
        self._require_openpyxl()
        path = self._chat_records_path()
        path.parent.mkdir(parents=True, exist_ok=True)

        if path.exists():
            wb = load_workbook(path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "records"
            ws.append(list(_XLSX_HEADERS))

        for m in messages:
            sn = (m.sender_name or "").strip()
            # 与 prepare_chat_records 列一致：无空格时当作 username，显示名仍用同串
            username_col = sn.lstrip("@") if sn and " " not in sn else ""
            display_col = sn or username_col
            ws.append(
                [
                    m.source_message_id,
                    m.message_date.strftime("%Y-%m-%d %H:%M:%S"),
                    m.sender_id if m.sender_id is not None else "",
                    username_col,
                    display_col,
                    m.text_content,
                ]
            )
        wb.save(path)

    def _load_script_messages_from_xlsx(
        self, limit: int | None
    ) -> list[ChatScriptMessageRecord]:
        cap = self._script_xlsx_row_cap(limit)
        records = self._load_messages_from_xlsx(cap)
        base = datetime.utcnow()
        step = max(1, self.settings.chat_script_row_interval_seconds)
        out: list[ChatScriptMessageRecord] = []
        cap_accounts = max(1, int(self.settings.max_active_accounts))
        for i, r in enumerate(records):
            out.append(
                ChatScriptMessageRecord(
                    script_id=1,
                    speaker_index=(i % cap_accounts) + 1,
                    message_time=base + timedelta(seconds=i * step),
                    text_content=r.text_content,
                )
            )
        return out

    async def _save_to_sqlite(self, messages: list[MessageRecord]) -> None:
        assert self.session_factory is not None
        rows = [
            {
                "source_group": m.source_group,
                "source_message_id": m.source_message_id,
                "sender_id": m.sender_id,
                "sender_name": m.sender_name,
                "text_content": m.text_content,
                "message_date": m.message_date,
            }
            for m in messages
        ]
        ins_m = sqlite_insert(CopiedMessage).values(rows)
        stmt = ins_m.on_conflict_do_update(
            index_elements=[
                CopiedMessage.source_group,
                CopiedMessage.source_message_id,
            ],
            set_={
                "text_content": ins_m.excluded.text_content,
                "sender_name": ins_m.excluded.sender_name,
                "message_date": ins_m.excluded.message_date,
            },
        )
        async with self.session_factory() as session:
            await session.execute(stmt)
            await session.commit()

    async def _save_to_redis(self, messages: list[MessageRecord]) -> None:
        assert self.redis_client is not None
        list_key = f"{self.settings.redis_key_prefix}:messages"
        dedup_key = f"{self.settings.redis_key_prefix}:seen"

        async with self.redis_client.pipeline(transaction=True) as pipe:
            for message in messages:
                unique_id = f"{message.source_group}:{message.source_message_id}"
                payload = asdict(message)
                payload["message_date"] = message.message_date.isoformat()
                pipe.sadd(dedup_key, unique_id)
                pipe.rpush(list_key, json.dumps(payload, ensure_ascii=False))
            await pipe.execute()

    async def _load_from_sqlite(self, limit: int) -> list[MessageRecord]:
        assert self.session_factory is not None
        async with self.session_factory() as session:
            stmt = (
                select(CopiedMessage)
                .order_by(CopiedMessage.message_date.asc())
                .limit(limit)
            )
            rows = (await session.execute(stmt)).scalars().all()
            return [
                MessageRecord(
                    source_group=row.source_group,
                    source_message_id=row.source_message_id,
                    sender_id=row.sender_id,
                    sender_name=row.sender_name,
                    text_content=row.text_content,
                    message_date=row.message_date,
                )
                for row in rows
            ]

    async def _load_from_redis(self, limit: int) -> list[MessageRecord]:
        assert self.redis_client is not None
        list_key = f"{self.settings.redis_key_prefix}:messages"
        items = await self.redis_client.lrange(list_key, 0, max(0, limit - 1))
        records: list[MessageRecord] = []
        for item in items:
            payload = json.loads(item)
            records.append(
                MessageRecord(
                    source_group=payload["source_group"],
                    source_message_id=payload["source_message_id"],
                    sender_id=payload.get("sender_id"),
                    sender_name=payload.get("sender_name"),
                    text_content=payload["text_content"],
                    message_date=datetime.fromisoformat(payload["message_date"]),
                )
            )
        return records
